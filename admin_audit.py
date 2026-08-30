import os
import threading
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import BASE_DIR
import database

logger = logging.getLogger("smartattend")

PRIVATE_DIR = BASE_DIR / "data" / "private"
EXCEL_REGISTRY_PATH = PRIVATE_DIR / "admin_registry.xlsx"

_excel_lock = threading.Lock()

HEADERS = [
    "Admin ID",
    "Full Name",
    "Username",
    "Email",
    "Registration Date",
    "Registration Time",
    "Last Login",
    "Account Status"
]

def _ensure_private_dir():
    try:
        PRIVATE_DIR.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        logger.error(f"Failed to create private audit directory: {e}")

def _apply_header_styling(sheet):
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for col_num, header in enumerate(HEADERS, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align
        cell.border = thin_border
        
    sheet.row_dimensions[1].height = 28

def _auto_adjust_column_widths(sheet):
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        sheet.column_dimensions[col_letter].width = max(max_len + 4, 14)

def init_admin_registry() -> bool:
    """
    Ensure the admin registry Excel file exists in data/private/admin_registry.xlsx.
    Populates existing admins from SQLite if the Excel file is being newly created.
    """
    _ensure_private_dir()
    with _excel_lock:
        try:
            if not EXCEL_REGISTRY_PATH.exists():
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.title = "Admin Registry"
                _apply_header_styling(sheet)
                
                # Pre-populate from SQLite users
                with database.db_session() as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT id, full_name, username, email, created_at FROM users ORDER BY id ASC;")
                    rows = cursor.fetchall()
                    
                thin_border = Border(
                    left=Side(style='thin', color='E2E8F0'),
                    right=Side(style='thin', color='E2E8F0'),
                    top=Side(style='thin', color='E2E8F0'),
                    bottom=Side(style='thin', color='E2E8F0')
                )
                
                for idx, row in enumerate(rows, start=2):
                    user_id = row["id"]
                    full_name = row["full_name"] or "Administrator"
                    username = row["username"]
                    email = row["email"] or f"{username}@smartattend.local"
                    created_at = row["created_at"] or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    try:
                        dt = datetime.strptime(created_at, "%Y-%m-%d %H:%M:%S")
                        reg_date = dt.strftime("%Y-%m-%d")
                        reg_time = dt.strftime("%H:%M:%S")
                    except Exception:
                        reg_date = created_at[:10] if len(created_at) >= 10 else created_at
                        reg_time = created_at[11:] if len(created_at) > 11 else ""
                        
                    data_row = [
                        user_id,
                        full_name,
                        username,
                        email,
                        reg_date,
                        reg_time,
                        "Never",
                        "Active"
                    ]
                    
                    for col_num, val in enumerate(data_row, 1):
                        c = sheet.cell(row=idx, column=col_num, value=val)
                        c.border = thin_border
                        if col_num in (1, 5, 6, 7, 8):
                            c.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            c.alignment = Alignment(horizontal="left", vertical="center")
                            
                _auto_adjust_column_widths(sheet)
                wb.save(str(EXCEL_REGISTRY_PATH))
                logger.info(f"Initialized admin registry Excel file at {EXCEL_REGISTRY_PATH}")
            return True
        except Exception as e:
            logger.error(f"Error initializing admin registry Excel file: {e}")
            return False

def record_admin_registration(
    admin_id: int,
    full_name: str,
    username: str,
    email: str,
    reg_datetime: Optional[str] = None,
    status: str = "Active"
) -> bool:
    """
    Append a newly registered administrator record to data/private/admin_registry.xlsx.
    Does NOT store passwords or password hashes.
    """
    _ensure_private_dir()
    now_dt = datetime.now()
    if reg_datetime:
        try:
            parsed = datetime.strptime(reg_datetime, "%Y-%m-%d %H:%M:%S")
            reg_date = parsed.strftime("%Y-%m-%d")
            reg_time = parsed.strftime("%H:%M:%S")
        except Exception:
            reg_date = now_dt.strftime("%Y-%m-%d")
            reg_time = now_dt.strftime("%H:%M:%S")
    else:
        reg_date = now_dt.strftime("%Y-%m-%d")
        reg_time = now_dt.strftime("%H:%M:%S")

    with _excel_lock:
        try:
            if not EXCEL_REGISTRY_PATH.exists():
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.title = "Admin Registry"
                _apply_header_styling(sheet)
            else:
                wb = openpyxl.load_workbook(str(EXCEL_REGISTRY_PATH))
                sheet = wb.active

            # Check if this admin ID or username already exists in Excel
            existing_row = None
            for row_idx in range(2, sheet.max_row + 1):
                cell_id = sheet.cell(row=row_idx, column=1).value
                cell_user = sheet.cell(row=row_idx, column=3).value
                if cell_id == admin_id or (cell_user and str(cell_user).lower() == username.lower()):
                    existing_row = row_idx
                    break

            target_row = existing_row if existing_row else (sheet.max_row + 1)
            
            thin_border = Border(
                left=Side(style='thin', color='E2E8F0'),
                right=Side(style='thin', color='E2E8F0'),
                top=Side(style='thin', color='E2E8F0'),
                bottom=Side(style='thin', color='E2E8F0')
            )

            data_row = [
                admin_id,
                full_name,
                username,
                email,
                reg_date,
                reg_time,
                "Never" if not existing_row else sheet.cell(row=target_row, column=7).value or "Never",
                status
            ]

            for col_num, val in enumerate(data_row, 1):
                c = sheet.cell(row=target_row, column=col_num, value=val)
                c.border = thin_border
                if col_num in (1, 5, 6, 7, 8):
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center")

            sheet.row_dimensions[target_row].height = 22
            _auto_adjust_column_widths(sheet)
            wb.save(str(EXCEL_REGISTRY_PATH))
            logger.info(f"Recorded admin registration in Excel audit: {username} (ID: {admin_id})")
            return True
        except Exception as e:
            logger.warning(f"Could not write to Excel admin registry (file may be open): {e}")
            return False

def record_admin_login(identifier: str, login_datetime: Optional[str] = None) -> bool:
    """
    Update the 'Last Login' timestamp for an administrator matching username or email.
    """
    if not identifier:
        return False
    _ensure_private_dir()
    now_str = login_datetime or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    clean_id = identifier.strip().lower()

    with _excel_lock:
        try:
            if not EXCEL_REGISTRY_PATH.exists():
                init_admin_registry()

            wb = openpyxl.load_workbook(str(EXCEL_REGISTRY_PATH))
            sheet = wb.active
            updated = False

            for row_idx in range(2, sheet.max_row + 1):
                username_val = str(sheet.cell(row=row_idx, column=3).value or '').strip().lower()
                email_val = str(sheet.cell(row=row_idx, column=4).value or '').strip().lower()

                if clean_id in (username_val, email_val):
                    sheet.cell(row=row_idx, column=7, value=now_str)
                    sheet.cell(row=row_idx, column=8, value="Active")
                    updated = True
                    break

            if updated:
                wb.save(str(EXCEL_REGISTRY_PATH))
                logger.info(f"Updated Last Login in Excel registry for '{identifier}' to {now_str}")
            return updated
        except Exception as e:
            logger.warning(f"Could not update Last Login in Excel registry (file may be open): {e}")
            return False

def get_all_audit_records() -> List[Dict[str, Any]]:
    """
    Read all administrator audit records from data/private/admin_registry.xlsx.
    Guarantees no passwords or hashes are returned.
    """
    init_admin_registry()
    records = []
    with _excel_lock:
        try:
            if not EXCEL_REGISTRY_PATH.exists():
                return []
            wb = openpyxl.load_workbook(str(EXCEL_REGISTRY_PATH), data_only=True)
            sheet = wb.active
            for row_idx in range(2, sheet.max_row + 1):
                admin_id = sheet.cell(row=row_idx, column=1).value
                if admin_id is None:
                    continue
                records.append({
                    "id": admin_id,
                    "full_name": sheet.cell(row=row_idx, column=2).value or "Administrator",
                    "username": sheet.cell(row=row_idx, column=3).value or "",
                    "email": sheet.cell(row=row_idx, column=4).value or "",
                    "reg_date": str(sheet.cell(row=row_idx, column=5).value or ""),
                    "reg_time": str(sheet.cell(row=row_idx, column=6).value or ""),
                    "last_login": str(sheet.cell(row=row_idx, column=7).value or "Never"),
                    "status": str(sheet.cell(row=row_idx, column=8).value or "Active")
                })
        except Exception as e:
            logger.error(f"Failed to read admin registry Excel: {e}")
            # Fallback to database user records if Excel file read fails
            with database.db_session() as conn:
                rows = conn.execute("SELECT id, full_name, username, email, created_at FROM users ORDER BY id ASC;").fetchall()
                for r in rows:
                    created_at = r["created_at"] or ""
                    records.append({
                        "id": r["id"],
                        "full_name": r["full_name"] or "Administrator",
                        "username": r["username"],
                        "email": r["email"] or f"{r['username']}@smartattend.local",
                        "reg_date": created_at[:10] if len(created_at) >= 10 else created_at,
                        "reg_time": created_at[11:] if len(created_at) > 11 else "",
                        "last_login": "Database Verified",
                        "status": "Active"
                    })
    return records
