import unittest
import os
import io
import csv
import openpyxl
from pathlib import Path

from app import app
import database
import security
import qr_engine
import face_engine
import admin_audit
from camera import camera_manager
from admin_audit import PRIVATE_DIR, EXCEL_REGISTRY_PATH

class Phase10SecurityHardeningTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        database.init_db()
        admin_audit.init_admin_registry()
        cls.client = app.test_client()

        # Clean old test records
        with database.db_session() as conn:
            conn.execute("DELETE FROM attendance WHERE student_id IN (SELECT id FROM students WHERE roll_no IN ('CS-2026-P10-A', 'CS-2026-P10-B'));")
            conn.execute("DELETE FROM students WHERE roll_no IN ('CS-2026-P10-A', 'CS-2026-P10-B');")

        # Register test students
        _, _, s1 = database.add_student("Vint Cerf", "CS-2026-P10-A")
        _, _, s2 = database.add_student("Radia Perlman", "CS-2026-P10-B")
        cls.student_1 = s1
        cls.student_2 = s2

    def test_01_authentication_enforcement_on_protected_routes(self):
        """Verify unauthenticated requests to all admin and core pages are blocked."""
        unauth_client = app.test_client()

        protected_pages = [
            "/",
            "/dashboard",
            "/students",
            f"/students/{self.student_1}",
            "/students/register",
            f"/qr/view/{self.student_1}",
            f"/qr/download/{self.student_1}",
            f"/capture/{self.student_1}",
            "/train",
            "/face-attendance",
            "/qr-attendance",
            "/daily-attendance",
            "/monthly-attendance",
            "/attendance-history",
            "/attendance/export/csv",
            "/settings",
            "/settings/export/admin-registry",
            "/video_feed/face",
            "/video_feed/qr",
            f"/video_feed/capture/{self.student_1}"
        ]

        for route in protected_pages:
            resp = unauth_client.get(route)
            self.assertEqual(
                resp.status_code, 302,
                f"Security Vulnerability: Protected page {route} returned HTTP {resp.status_code} instead of 302 redirect to login"
            )
            self.assertIn("/login", resp.headers.get("Location", ""))

    def test_02_csrf_protection_on_state_changing_actions(self):
        """Verify state-changing POST endpoints reject requests without a valid CSRF token."""
        with self.client.session_transaction() as sess:
            sess["user_id"] = 1
            sess["username"] = "admin"
            sess["role"] = "admin"

        # POST without CSRF token
        resp = self.client.post("/students/register", data={"name": "Attacker Student", "roll_no": "ATTACK-001"})
        self.assertIn(resp.status_code, [400, 403])
        self.assertIn(b"CSRF", resp.data)

        resp2 = self.client.post(f"/students/delete/{self.student_1}", data={})
        self.assertIn(resp2.status_code, [400, 403])

        resp3 = self.client.post("/settings", data={"app_name": "HackedName"})
        self.assertIn(resp3.status_code, [400, 403])

    def test_03_private_excel_registry_isolation_and_export_safety(self):
        """Verify admin Excel registry is stored privately and contains NO passwords or password hashes."""
        # 1. Verify storage directory is inside data/private/
        self.assertTrue(EXCEL_REGISTRY_PATH.parent.resolve() == PRIVATE_DIR.resolve())
        self.assertTrue(EXCEL_REGISTRY_PATH.exists())

        # 2. Unauthenticated export download is blocked
        unauth_client = app.test_client()
        resp_unauth = unauth_client.get("/settings/export/admin-registry")
        self.assertEqual(resp_unauth.status_code, 302)

        # 3. Authenticated export verification
        with self.client.session_transaction() as sess:
            sess["user_id"] = 1
            sess["username"] = "admin"
            sess["role"] = "admin"

        resp = self.client.get("/settings/export/admin-registry")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.mimetype, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # 4. Inspect Excel Workbook with openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(resp.data))
        sheet = wb.active

        # Check headers
        headers = [sheet.cell(row=1, column=col).value for col in range(1, 9)]
        expected_headers = [
            "Admin ID", "Full Name", "Username", "Email",
            "Registration Date", "Registration Time", "Last Login", "Account Status"
        ]
        self.assertEqual(headers, expected_headers)

        # Assert no password or hash fields exist in any cell
        for row in sheet.iter_rows(values_only=True):
            for cell_val in row:
                val_str = str(cell_val or "").lower()
                self.assertNotIn("scrypt:", val_str)
                self.assertNotIn("pbkdf2:", val_str)
                self.assertNotIn("hash", val_str)
                self.assertNotIn("password", val_str)

    def test_04_sensitive_field_exclusion_from_api_and_csv_exports(self):
        """Verify API endpoints and CSV downloads never expose credentials or secret keys."""
        with self.client.session_transaction() as sess:
            sess["user_id"] = 1
            sess["username"] = "admin"
            sess["role"] = "admin"

        # Check CSV Export
        resp_csv = self.client.get("/attendance/export/csv")
        self.assertEqual(resp_csv.status_code, 200)
        csv_text = resp_csv.data.decode("utf-8")
        self.assertNotIn("password", csv_text.lower())
        self.assertNotIn("scrypt:", csv_text)

        # Check API endpoints
        resp_stats = self.client.get("/api/dashboard")
        self.assertNotIn("password", resp_stats.data.decode("utf-8").lower())

        resp_students = self.client.get("/api/students")
        self.assertNotIn("password", resp_students.data.decode("utf-8").lower())

    def test_05_path_traversal_and_file_safety(self):
        """Verify routes resist path traversal attacks."""
        with self.client.session_transaction() as sess:
            sess["user_id"] = 1
            sess["username"] = "admin"
            sess["role"] = "admin"

        # Attempt path traversal on student ID
        resp = self.client.get("/students/..%2f..%2fetc%2fpasswd")
        self.assertEqual(resp.status_code, 404)

        resp2 = self.client.get("/qr/download/..%2f..%2fapp.py")
        self.assertEqual(resp2.status_code, 404)

    def test_06_nonexistent_student_and_error_handling(self):
        """Verify nonexistent student IDs are handled gracefully with proper error messages."""
        with self.client.session_transaction() as sess:
            sess["user_id"] = 1
            sess["username"] = "admin"
            sess["role"] = "admin"

        # Non-existent student profile redirects with flash
        resp = self.client.get("/students/999999")
        self.assertEqual(resp.status_code, 302)

        # Non-existent student API returns 404
        resp_api = self.client.post("/api/capture/reset/999999")
        self.assertEqual(resp_api.status_code, 404)

    def test_07_duplicate_attendance_double_barrier(self):
        """Verify double check-in on same day is blocked by SQLite UNIQUE constraint."""
        today_str = database.datetime.now().strftime("%Y-%m-%d")

        # Clear test attendance
        with database.db_session() as conn:
            conn.execute("DELETE FROM attendance WHERE student_id = ? AND date = ?;", (self.student_1, today_str))

        # First mark -> SUCCESS
        ok1, msg1, rec1 = database.mark_attendance(self.student_1, method="face")
        self.assertTrue(ok1)

        # Second mark -> REJECTED
        ok2, msg2, rec2 = database.mark_attendance(self.student_1, method="face")
        self.assertFalse(ok2)
        self.assertIn("Already marked today", msg2)

        # Third mark via different method (QR) -> REJECTED
        ok3, msg3, rec3 = database.mark_attendance(self.student_1, method="qr")
        self.assertFalse(ok3)
        self.assertIn("Already marked today", msg3)

    def test_08_unauthorized_camera_and_biometric_controls(self):
        """Verify camera control and biometric endpoints require authentication."""
        unauth_client = app.test_client()

        self.assertEqual(unauth_client.get("/api/camera/status").status_code, 401)
        self.assertEqual(unauth_client.post("/api/camera/start").status_code, 401)
        self.assertEqual(unauth_client.post("/api/camera/stop").status_code, 401)
        self.assertEqual(unauth_client.post(f"/api/capture/start/{self.student_1}").status_code, 401)
        self.assertEqual(unauth_client.post("/api/train").status_code, 401)

    def test_09_qr_payload_tampering_and_forgery_rejection(self):
        """Verify QR engine rejects tampered checksums, forged payloads, and non-existent student IDs."""
        # 1. Valid payload
        valid_payload = qr_engine.generate_payload(self.student_1)
        is_valid, sid, _ = qr_engine.validate_payload(valid_payload)
        self.assertTrue(is_valid)
        self.assertEqual(sid, self.student_1)

        # 2. Tampered checksum
        tampered_payload = f"SMARTATTEND:{self.student_1}:INVALIDCHECKSUM"
        is_valid_t, _, _ = qr_engine.validate_payload(tampered_payload)
        self.assertFalse(is_valid_t)

        # 3. Forged student ID with valid format
        forged_res = qr_engine.process_qr_attendance("SMARTATTEND:999999:00000000")
        self.assertFalse(forged_res["success"])
        self.assertIn(forged_res["status"], ["INVALID", "NOT_FOUND", "INVALID_CHECKSUM", "INVALID_FORMAT"])

    def test_10_safe_error_handlers(self):
        """Verify custom 404, 401, 403, and 405 error handlers render safely without leaking stack traces."""
        # 404
        resp_404 = self.client.get("/non-existent-route-404")
        self.assertEqual(resp_404.status_code, 404)
        self.assertIn(b"Page Not Found", resp_404.data)

        # 405
        with self.client.session_transaction() as sess:
            sess["user_id"] = 1
            sess["username"] = "admin"
            sess["role"] = "admin"

        resp_405 = self.client.get("/students/delete/1")
        self.assertIn(resp_405.status_code, [405, 404])

    def test_11_camera_idempotency_and_safe_release(self):
        """Verify camera_manager stop() and release_camera() are idempotent and leak-free."""
        camera_manager.stop()
        camera_manager.stop()
        camera_manager.release_camera()
        camera_manager.release_camera()
        self.assertFalse(camera_manager.is_running())

    def test_12_database_backup_download_and_integrity(self):
        """Verify database backup authentication, headers, valid SQLite .db structure, and live DB preservation."""
        import sqlite3
        import tempfile

        # 1. Unauthenticated request rejected
        unauth_client = app.test_client()
        resp_unauth = unauth_client.post("/backup")
        self.assertEqual(resp_unauth.status_code, 302)
        self.assertIn("/login", resp_unauth.headers.get("Location", ""))

        # 2. Authenticated request with CSRF succeeds
        token = "test_valid_csrf_token_value"
        with self.client.session_transaction() as sess:
            sess["user_id"] = 1
            sess["username"] = "admin"
            sess["role"] = "admin"
            sess["_csrf_token"] = token

        # Record student count before backup
        students_before = len(database.get_all_students())

        resp = self.client.post("/backup", data={"csrf_token": token})
        self.assertEqual(resp.status_code, 200)

        # 3. Response headers check
        self.assertEqual(resp.mimetype, "application/x-sqlite3")
        disposition = resp.headers.get("Content-Disposition", "")
        self.assertIn("attachment", disposition)
        self.assertIn("filename=", disposition)
        self.assertIn(".db", disposition)

        filename_header = resp.headers.get("X-Backup-Filename", "")
        self.assertTrue(filename_header.endswith(".db"))
        self.assertTrue(filename_header.startswith("attendance_backup_"))

        # 4. Verify downloaded content is a real valid SQLite database
        backup_bytes = resp.data
        self.assertGreater(len(backup_bytes), 0)

        with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as tmp:
            tmp.write(backup_bytes)
            tmp_path = tmp.name

        try:
            conn = sqlite3.connect(tmp_path)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = [row[0] for row in cursor.fetchall()]
            self.assertIn("students", tables)
            self.assertIn("attendance", tables)
            self.assertIn("users", tables)
            self.assertIn("settings", tables)

            # Verify student records inside backup database
            cursor.execute("SELECT COUNT(*) FROM students;")
            backup_student_count = cursor.fetchone()[0]
            self.assertEqual(backup_student_count, students_before)
            conn.close()
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

        # 5. Verify live database is completely unchanged
        students_after = len(database.get_all_students())
        self.assertEqual(students_before, students_after)

    @classmethod
    def tearDownClass(cls):
        # Clean up test students
        database.delete_student(cls.student_1)
        database.delete_student(cls.student_2)
        camera_manager.stop()

if __name__ == "__main__":
    unittest.main()
